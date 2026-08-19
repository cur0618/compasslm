import unittest
from tempfile import NamedTemporaryFile

from src.table_facts import build_definition_facts_from_lines, build_flat_table_row_fact_text, build_table_row_fact_text
from src.utils import chunk_txt_items, chunk_xlsx_rows, load_xlsx


class TableFactTests(unittest.TestCase):
    def test_build_table_row_fact_includes_subject_alias_and_fields(self):
        fact = build_table_row_fact_text(
            headers=["답례품", "조사명", "지급단가", "지급횟수"],
            values=["지류, 현금", "농·어가경제조사", "40", "12"],
        )

        self.assertIn("표의미: kind=table_row", fact)
        self.assertIn("subject=농·어가경제조사", fact)
        self.assertIn("aliases=농·어가경제조사, 농어가경제조사, 농가경제조사, 어가경제조사", fact)
        self.assertIn("지급단가=40천원", fact)
        self.assertIn("지급횟수=12", fact)

    def test_flat_table_fact_keeps_extra_name_tokens_with_subject(self):
        fact = build_table_row_fact_text(
            headers=["조사명", "지급단가", "지급횟수"],
            values=["경제활동인구조사 본조사", "10", "12"],
        )
        flat_fact = build_flat_table_row_fact_text(
            "조사명 지급단가 지급횟수",
            "경제활동인구조사 본조사 10 12",
        )

        self.assertEqual(flat_fact, fact)
        self.assertIn("subject=경제활동인구조사 본조사", flat_fact)
        self.assertIn("지급단가=10천원", flat_fact)

    def test_definition_like_guide_lines_are_grouped_as_definition_facts(self):
        lines = [
            "논",
            "① 일모작",
            "물을 이용하여 논벼, 미나리, 연, 왕골 등의 작물을 주로 재배하는 농지",
            "동일한 경작지에서 한 해에 한 차례만 작물을 거두는 일",
            "② 이모작",
            "동일한 경작지에서 한 해에 두 차례 다른 작물을 거두는 일",
            "속 청",
            "농가에서 통상 부르는 경지의 명칭 및 소재지 등을 기입한다.",
        ]

        facts = build_definition_facts_from_lines(lines)

        joined = "\n".join(facts)
        self.assertIn("표의미: kind=definition_block | subject=논 > ① 일모작", joined)
        self.assertIn("동일한 경작지에서 한 해에 한 차례만 작물을 거두는 일", joined)
        self.assertIn("표의미: kind=definition_block | subject=논 > ② 이모작", joined)
        self.assertIn("표의미: kind=definition_block | subject=속 청", joined)

    def test_definition_facts_filter_page_chrome_and_toc_noise(self):
        lines = [
            "2024년 농가경제조사 지침서",
            "Ⅰ",
            "조 사 개 요",
            "Ⅱ",
            "현 장 조 사",
            "header",
            "STATISTICS KOREA",
            "image",
            "imgs/img_in_image_box_139_180_296_273.jpg",
            "논",
            "① 일모작",
            "물을 이용하여 논벼, 미나리, 연, 왕골 등의 작물을 주로 재배하는 농지",
            "동일한 경작지에서 한 해에 한 차례만 작물을 거두는 일",
            "속 청",
            "농가에서 통상 부르는 경지의 명칭 및 소재지 등을 기입한다.",
        ]

        facts = build_definition_facts_from_lines(lines)

        joined = "\n".join(facts)
        self.assertIn("표의미: kind=definition_block | subject=논 > ① 일모작", joined)
        self.assertIn("표의미: kind=definition_block | subject=속 청", joined)
        self.assertNotIn("subject=Ⅰ", joined)
        self.assertNotIn("subject=Ⅱ", joined)
        self.assertNotIn("subject=header", joined)
        self.assertNotIn("subject=image", joined)
        self.assertNotIn("STATISTICS KOREA", joined)

    def test_definition_facts_filter_single_character_toc_trails_and_garbled_subjects(self):
        lines = [
            "요",
            "표",
            "검",
            "력",
            "목 차",
            "차례",
            "··········",
            "포도 판매",
            "농가에서 생산한 포도를 원물로 판매한 경우 판매량과 금액을 기입한다.",
        ]

        facts = build_definition_facts_from_lines(lines)

        joined = "\n".join(facts)
        self.assertIn("subject=포도 판매", joined)
        self.assertNotIn("subject=요", joined)
        self.assertNotIn("subject=표", joined)
        self.assertNotIn("subject=검", joined)
        self.assertNotIn("subject=력", joined)
        self.assertNotIn("목 차", joined)

    def test_txt_chunks_include_definition_facts_for_guide_like_lines(self):
        chunks = chunk_txt_items(
            [
                {"text": "논", "line_start": 1, "line_end": 1, "file_path": "guide.txt", "is_section": False},
                {"text": "① 일모작", "line_start": 2, "line_end": 2, "file_path": "guide.txt", "is_section": False},
                {
                    "text": "동일한 경작지에서 한 해에 한 차례만 작물을 거두는 일",
                    "line_start": 3,
                    "line_end": 3,
                    "file_path": "guide.txt",
                    "is_section": False,
                },
            ],
            target_tokens=1,
            min_tokens=1,
            max_tokens=80,
        )

        self.assertTrue(
            any("표의미: kind=definition_block | subject=논 > ① 일모작" in chunk["text"] for chunk in chunks)
        )

    def test_xlsx_chunks_include_table_row_fact_text(self):
        try:
            import openpyxl
        except Exception:
            self.skipTest("openpyxl unavailable")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "기준"
        ws.append(["조사명", "지급단가", "지급횟수"])
        ws.append(["농어가경제조사", 40, 12])
        with NamedTemporaryFile(suffix=".xlsx") as tmp:
            wb.save(tmp.name)
            rows = load_xlsx(tmp.name)

        chunks = chunk_xlsx_rows(rows, group_min_rows=1, group_max_rows=1)

        self.assertIn("표의미: kind=table_row", chunks[0]["text"])
        self.assertIn("subject=농어가경제조사", chunks[0]["text"])
        self.assertIn("지급단가=40천원", chunks[0]["text"])

    def test_xlsx_loader_emits_structure_metadata_for_rows(self):
        try:
            import openpyxl
        except Exception:
            self.skipTest("openpyxl unavailable")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "지급기준"
        ws.append(["조사명", "지급단가"])
        ws.append(["농가경제조사", 40])
        with NamedTemporaryFile(suffix=".xlsx") as tmp:
            wb.save(tmp.name)
            rows = load_xlsx(tmp.name, structure_v2=True)

        self.assertEqual(rows[0]["chunk_kind"], "table_row")
        self.assertEqual(rows[0]["heading_path"], ["지급기준"])
        self.assertEqual(rows[0]["table_id"], "지급기준:table:1")
        self.assertEqual(rows[0]["row_no"], 2)
        self.assertEqual(rows[0]["parent_chunk_key"], "xlsx:지급기준:table:1:row:2")
        self.assertIn("source=", rows[0]["embedding_text"])
        self.assertIn("path=지급기준", rows[0]["embedding_text"])

    def test_xlsx_structure_loader_separates_multiple_tables_in_one_sheet(self):
        try:
            import openpyxl
        except Exception:
            self.skipTest("openpyxl unavailable")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "운영기준"
        ws.append(["답례품 지급표"])
        ws.append(["조사명", "지급단가"])
        ws.append(["농가경제조사", 40])
        ws.append([])
        ws.append(["조사 일정표"])
        ws.append(["조사명", "기준월"])
        ws.append(["농가경제조사", "4월"])
        with NamedTemporaryFile(suffix=".xlsx") as tmp:
            wb.save(tmp.name)
            rows = load_xlsx(tmp.name, structure_v2=True)

        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[0]["table_id"], "운영기준:table:1")
        self.assertEqual(rows[1]["table_id"], "운영기준:table:2")
        self.assertEqual(rows[0]["heading_path"], ["운영기준", "답례품 지급표"])
        self.assertEqual(rows[1]["heading_path"], ["운영기준", "조사 일정표"])
        self.assertIn("지급단가=40", rows[0]["text"])
        self.assertIn("기준월=4월", rows[1]["text"])


if __name__ == "__main__":
    unittest.main()
