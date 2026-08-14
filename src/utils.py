import os
import re
import time
from collections import defaultdict
from typing import Any, Dict, List, Optional, Tuple

import chardet
import openpyxl

from src.table_facts import build_definition_facts_from_lines, build_table_row_fact_text
from src.document_structure import normalize_structure_record


SECTION_LINE_PATTERN = re.compile(
    r"^(#{1,6}\s+.+|(?:\d+(?:\.\d+){0,3}|[가-힣A-Za-z]\))\s+.+|제\s*\d+\s*(장|절|항|조).+)$"
)
HEADER_FOOTER_PATTERNS = (
    re.compile(r"^page\s*\d+\s*(of\s*\d+)?$", re.IGNORECASE),
    re.compile(r"^\d+\s*/\s*\d+$"),
    re.compile(r"^[-_=]{3,}$"),
)
TOKEN_PATTERN = re.compile(r"[0-9A-Za-z가-힣]+")
WHITESPACE_PATTERN = re.compile(r"\s+")


def detect_encoding(file_path: str) -> str:
    """Best-effort encoding detection with deterministic fallbacks."""
    with open(file_path, "rb") as f:
        sample = f.read(65536)

    detected = (chardet.detect(sample or b"").get("encoding") or "").strip().lower()
    if detected in {"utf-8", "utf_8", "utf-8-sig", "cp949", "euc-kr"}:
        return detected
    return "utf-8-sig"


def _read_text_with_fallback(file_path: str) -> str:
    detected = detect_encoding(file_path)
    tried: List[str] = []
    for enc in [detected, "utf-8-sig", "utf-8", "cp949", "euc-kr"]:
        if enc in tried:
            continue
        tried.append(enc)
        try:
            with open(file_path, "r", encoding=enc) as f:
                return f.read()
        except Exception:
            continue

    # Last-resort decode for malformed files.
    with open(file_path, "rb") as f:
        raw = f.read()
    return raw.decode("utf-8", errors="replace")


def _normalize_line(text: str) -> str:
    return WHITESPACE_PATTERN.sub(" ", (text or "").strip())


def _is_probable_header_footer(line: str) -> bool:
    if not line:
        return True
    return any(pat.match(line) for pat in HEADER_FOOTER_PATTERNS)


def _is_section_line(line: str) -> bool:
    if not line:
        return False
    return bool(SECTION_LINE_PATTERN.match(line))


def _estimate_tokens(text: str) -> int:
    # Lightweight heuristic for chunk sizing.
    if not text:
        return 0
    tokens = TOKEN_PATTERN.findall(text)
    return max(1, int(len(tokens) * 1.08))


def load_txt(file_path: str) -> List[Dict[str, Any]]:
    """
    Load TXT while preserving section markers and line numbers.
    Returns normalized non-empty lines with metadata.
    """
    content = _read_text_with_fallback(file_path)
    filename = os.path.basename(file_path)

    rows: List[Dict[str, Any]] = []
    for idx, raw in enumerate(content.splitlines(), start=1):
        normalized = _normalize_line(raw)
        if not normalized:
            continue

        section_line = _is_section_line(normalized)
        if _is_probable_header_footer(normalized) and not section_line:
            continue

        rows.append(
            {
                "text": normalized,
                "line_start": idx,
                "line_end": idx,
                "file_path": filename,
                "is_section": section_line,
            }
        )

    return rows


def _excel_col_name(index: int) -> str:
    result = ""
    i = max(1, int(index))
    while i > 0:
        i, rem = divmod(i - 1, 26)
        result = chr(ord("A") + rem) + result
    return result


def _normalize_cell_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return _normalize_line(value)
    return _normalize_line(str(value))


def _build_merged_lookup(sheet) -> Dict[Tuple[int, int], Any]:
    """Map every merged cell coordinate to its top-left value."""
    lookup: Dict[Tuple[int, int], Any] = {}
    for merged in sheet.merged_cells.ranges:
        min_row, min_col, max_row, max_col = merged.min_row, merged.min_col, merged.max_row, merged.max_col
        top_left_value = sheet.cell(min_row, min_col).value
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                lookup[(r, c)] = top_left_value
    return lookup


def _pick_header_row(sheet) -> int:
    max_scan = min(sheet.max_row, 5)
    for r in range(1, max_scan + 1):
        values = [_normalize_cell_value(sheet.cell(r, c).value) for c in range(1, sheet.max_column + 1)]
        non_empty = [v for v in values if v]
        if len(non_empty) >= max(2, int(sheet.max_column * 0.3)):
            return r
    return 1


def _extract_headers(sheet, header_row: int) -> List[str]:
    headers: List[str] = []
    for c in range(1, sheet.max_column + 1):
        raw = _normalize_cell_value(sheet.cell(header_row, c).value)
        if raw:
            headers.append(raw)
        else:
            headers.append(f"Col{_excel_col_name(c)}")
    return headers


def _load_xlsx_structure_records(
    sheet,
    *,
    filename: str,
    merged_policy: str,
    note_policy: str,
) -> List[Dict[str, Any]]:
    merged_lookup = _build_merged_lookup(sheet) if merged_policy == "expand" else {}
    segments: List[List[int]] = []
    current: List[int] = []
    for row_idx in range(1, sheet.max_row + 1):
        values = [
            _normalize_cell_value(
                sheet.cell(row_idx, col_idx).value
                if sheet.cell(row_idx, col_idx).value not in (None, "")
                else merged_lookup.get((row_idx, col_idx))
            )
            for col_idx in range(1, sheet.max_column + 1)
        ]
        if any(values):
            current.append(row_idx)
        elif current:
            segments.append(current)
            current = []
    if current:
        segments.append(current)

    records: List[Dict[str, Any]] = []
    table_no = 0
    for segment in segments:
        header_row = 0
        for row_idx in segment:
            non_empty = sum(
                1
                for col_idx in range(1, sheet.max_column + 1)
                if _normalize_cell_value(
                    sheet.cell(row_idx, col_idx).value
                    if sheet.cell(row_idx, col_idx).value not in (None, "")
                    else merged_lookup.get((row_idx, col_idx))
                )
            )
            if non_empty >= 2:
                header_row = row_idx
                break
        if not header_row:
            continue
        table_no += 1
        headers = []
        for col_idx in range(1, sheet.max_column + 1):
            raw = sheet.cell(header_row, col_idx).value
            if raw in (None, ""):
                raw = merged_lookup.get((header_row, col_idx))
            headers.append(_normalize_cell_value(raw) or f"Col{_excel_col_name(col_idx)}")
        title_parts: List[str] = []
        for title_row in segment:
            if title_row >= header_row:
                break
            for col_idx in range(1, sheet.max_column + 1):
                value = _normalize_cell_value(sheet.cell(title_row, col_idx).value)
                if value and value not in title_parts:
                    title_parts.append(value)
        title = " ".join(title_parts)
        table_id = f"{sheet.title}:table:{table_no}"
        heading_path = [sheet.title] + ([title] if title else [])

        for row_idx in segment:
            if row_idx <= header_row:
                continue
            values: List[Tuple[str, str]] = []
            non_empty_count = 0
            for col_idx, col_name in enumerate(headers, start=1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                raw_val = cell.value
                if raw_val in (None, "") and merged_policy == "expand":
                    raw_val = merged_lookup.get((row_idx, col_idx), raw_val)
                value = _normalize_cell_value(raw_val)
                if note_policy == "footnote" and cell.comment and cell.comment.text:
                    note = _normalize_line(cell.comment.text)
                    if note:
                        value = f"{value} [주석:{note}]" if value else f"[주석:{note}]"
                if value:
                    non_empty_count += 1
                values.append((col_name, value))
            if non_empty_count == 0:
                continue
            row_serial = " | ".join(f"{key}={value}" for key, value in values if value)
            record = {
                "file_path": filename,
                "sheet": sheet.title,
                "row": row_idx,
                "row_end": row_idx,
                "columns": list(headers),
                "row_text": f"Row {row_idx}: {row_serial}",
                "text": f"Row {row_idx}: {row_serial}",
                "fact_text": build_table_row_fact_text(headers, [value for _, value in values]),
                "heading_path": heading_path,
                "chunk_kind": "table_row",
                "table_id": table_id,
                "row_no": row_idx,
                "cell_no": 0,
                "structure_path": f"sheet:{sheet.title}/table:{table_no}/row:{row_idx}",
                "parent_chunk_key": f"xlsx:{sheet.title}:table:{table_no}:row:{row_idx}",
                "is_derived": False,
            }
            records.append(normalize_structure_record(record, source_type="xlsx", doc_role=""))
    return records


def load_xlsx(
    file_path: str,
    merged_cell_policy: Optional[str] = None,
    comment_policy: Optional[str] = None,
    structure_v2: bool = False,
) -> List[Dict[str, Any]]:
    """
    XLSX -> row records for later grouped chunking.

    Serialization target:
    [Sheet: ...]
    [Columns: ...]
    Row N: col=value | ...
    """
    merged_policy = (merged_cell_policy or os.getenv("XLSX_MERGED_CELL_POLICY", "expand")).strip().lower()
    note_policy = (comment_policy or os.getenv("XLSX_COMMENT_POLICY", "footnote")).strip().lower()
    max_sheets = max(1, int(os.getenv("XLSX_MAX_SHEETS", "20")))
    max_rows = max(1, int(os.getenv("XLSX_MAX_ROWS", "50000")))
    max_cols = max(1, int(os.getenv("XLSX_MAX_COLS", "200")))
    parse_timeout_sec = max(1.0, float(os.getenv("XLSX_PARSE_TIMEOUT_SEC", "25")))
    started = time.monotonic()

    def _check_timeout():
        if (time.monotonic() - started) > parse_timeout_sec:
            raise TimeoutError(f"XLSX parsing timeout exceeded ({parse_timeout_sec:.1f}s).")

    records: List[Dict[str, Any]] = []
    wb = openpyxl.load_workbook(file_path, data_only=True)
    filename = os.path.basename(file_path)
    _check_timeout()

    if len(wb.sheetnames) > max_sheets:
        raise ValueError(f"Too many sheets in XLSX ({len(wb.sheetnames)} > {max_sheets}).")

    for sheet_name in wb.sheetnames:
        _check_timeout()
        sheet = wb[sheet_name]
        if sheet.max_row <= 0 or sheet.max_column <= 0:
            continue
        if sheet.max_row > max_rows:
            raise ValueError(f"Too many rows in sheet '{sheet_name}' ({sheet.max_row} > {max_rows}).")
        if sheet.max_column > max_cols:
            raise ValueError(f"Too many columns in sheet '{sheet_name}' ({sheet.max_column} > {max_cols}).")

        if structure_v2:
            records.extend(
                _load_xlsx_structure_records(
                    sheet,
                    filename=filename,
                    merged_policy=merged_policy,
                    note_policy=note_policy,
                )
            )
            continue

        header_row = _pick_header_row(sheet)
        headers = _extract_headers(sheet, header_row)
        merged_lookup = _build_merged_lookup(sheet) if merged_policy == "expand" else {}

        for row_idx in range(header_row + 1, sheet.max_row + 1):
            if row_idx % 256 == 0:
                _check_timeout()
            values: List[Tuple[str, str]] = []
            non_empty_count = 0

            for col_idx, col_name in enumerate(headers, start=1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                raw_val = cell.value
                if (raw_val is None or raw_val == "") and merged_policy == "expand":
                    raw_val = merged_lookup.get((row_idx, col_idx), raw_val)

                val = _normalize_cell_value(raw_val)

                if note_policy == "footnote" and cell.comment and cell.comment.text:
                    note = _normalize_line(cell.comment.text)
                    if note:
                        if val:
                            val = f"{val} [주석:{note}]"
                        else:
                            val = f"[주석:{note}]"

                if val:
                    non_empty_count += 1
                values.append((col_name, val))

            if non_empty_count == 0:
                continue

            row_serial = " | ".join(f"{k}={v}" for k, v in values if k)
            record = {
                    "file_path": filename,
                    "sheet": sheet_name,
                    "row": row_idx,
                    "columns": list(headers),
                    "row_text": f"Row {row_idx}: {row_serial}",
                    "fact_text": build_table_row_fact_text(headers, [value for _, value in values]),
                }
            if structure_v2:
                table_id = f"{sheet_name}:table:1"
                record.update(
                    {
                        "text": record["row_text"],
                        "heading_path": [sheet_name],
                        "chunk_kind": "table_row",
                        "table_id": table_id,
                        "row_no": row_idx,
                        "cell_no": 0,
                        "structure_path": f"sheet:{sheet_name}/table:1/row:{row_idx}",
                        "parent_chunk_key": f"xlsx:{sheet_name}:table:1:row:{row_idx}",
                        "is_derived": False,
                    }
                )
                record = normalize_structure_record(record, source_type="xlsx", doc_role="")
            records.append(record)

    return records


def _find_previous_section(lines: List[Dict[str, Any]], idx: int, lookback: int = 36) -> str:
    start = max(0, idx - lookback)
    for i in range(idx - 1, start - 1, -1):
        if lines[i].get("is_section"):
            return lines[i].get("text", "")
    return ""


def chunk_txt_items(
    lines: List[Dict[str, Any]],
    target_tokens: int = 620,
    min_tokens: int = 420,
    max_tokens: int = 780,
    overlap_ratio: float = 0.15,
) -> List[Dict[str, Any]]:
    """Chunk TXT lines with section retention and token overlap."""
    if not lines:
        return []

    chunks: List[Dict[str, Any]] = []
    global_definition_facts = build_definition_facts_from_lines([item.get("text", "") for item in lines])
    start = 0
    n = len(lines)
    overlap_tokens = max(30, int(max_tokens * max(0.1, min(0.2, overlap_ratio))))

    while start < n:
        end = start
        used_tokens = 0

        while end < n:
            line_text = lines[end]["text"]
            add_tokens = _estimate_tokens(line_text)
            if end > start and used_tokens + add_tokens > max_tokens:
                break
            used_tokens += add_tokens
            end += 1

            if used_tokens >= target_tokens and used_tokens >= min_tokens:
                break

        while end < n and used_tokens < min_tokens:
            add_tokens = _estimate_tokens(lines[end]["text"])
            if end > start and used_tokens + add_tokens > max_tokens:
                break
            used_tokens += add_tokens
            end += 1

        window = lines[start:end]
        if not window:
            start += 1
            continue

        has_section = any(item.get("is_section") for item in window)
        section_hint = _find_previous_section(lines, start)
        combined_text = "\n".join(item.get("text", "") for item in window)

        if section_hint and not has_section:
            combined_text = f"[Section: {section_hint}]\n{combined_text}"
        chunks.append(
            {
                "text": combined_text,
                "file_path": window[0].get("file_path", ""),
                "line_start": int(window[0].get("line_start", 0) or 0),
                "line_end": int(window[-1].get("line_end", 0) or 0),
                "section": section_hint if section_hint else "",
            }
        )

        if end >= n:
            break

        back_tokens = 0
        next_start = end
        while next_start > start and back_tokens < overlap_tokens:
            next_start -= 1
            back_tokens += _estimate_tokens(lines[next_start].get("text", ""))

        if next_start <= start:
            start = end
        else:
            start = next_start

    if global_definition_facts:
        chunks.append(
            {
                "text": "\n".join(global_definition_facts),
                "file_path": lines[0].get("file_path", ""),
                "line_start": int(lines[0].get("line_start", 0) or 0),
                "line_end": int(lines[-1].get("line_end", 0) or 0),
                "section": "DefinitionFacts",
            }
        )

    return chunks


def chunk_xlsx_rows(
    rows: List[Dict[str, Any]],
    group_min_rows: int = 10,
    group_max_rows: int = 30,
    overlap_rows: int = 3,
    target_tokens: int = 620,
    max_tokens: int = 780,
) -> List[Dict[str, Any]]:
    """Group XLSX rows per sheet to avoid semantic fragmentation."""
    if not rows:
        return []

    grouped: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for item in rows:
        grouped[item.get("sheet", "Sheet1")].append(item)

    chunks: List[Dict[str, Any]] = []

    for sheet_name, sheet_rows in grouped.items():
        sheet_rows.sort(key=lambda x: int(x.get("row", 0) or 0))
        start = 0
        n = len(sheet_rows)

        while start < n:
            end = start
            row_count = 0
            used_tokens = 0

            while end < n and row_count < group_max_rows:
                row_line = sheet_rows[end].get("row_text", "")
                add_tokens = _estimate_tokens(row_line)
                if row_count >= group_min_rows and used_tokens + add_tokens > max_tokens:
                    break
                used_tokens += add_tokens
                row_count += 1
                end += 1
                if row_count >= group_min_rows and used_tokens >= target_tokens:
                    break

            if end == start:
                end = min(n, start + 1)

            window = sheet_rows[start:end]
            if not window:
                start += 1
                continue

            columns = window[0].get("columns", []) or []
            columns_line = "|".join(str(col) for col in columns)
            text_lines = [
                f"[Sheet: {sheet_name}]",
                f"[Columns: {columns_line}]",
            ]
            text_lines.extend(item.get("row_text", "") for item in window)
            text_lines.extend(item.get("fact_text", "") for item in window if item.get("fact_text"))
            text = "\n".join(text_lines)

            row_start = int(window[0].get("row", 0) or 0)
            row_end = int(window[-1].get("row", 0) or 0)
            chunks.append(
                {
                    "text": text,
                    "file_path": window[0].get("file_path", ""),
                    "sheet": sheet_name,
                    "row": row_start,
                    "row_end": row_end,
                    "section": f"Sheet:{sheet_name}",
                }
            )

            if end >= n:
                break

            next_start = max(start + 1, end - max(1, overlap_rows))
            start = next_start

    return chunks


def chunk_text(
    lines: List[Dict[str, Any]],
    max_chars: int = 1200,
    overlap: int = 180,
    min_chunk_chars: int = 280,
) -> List[Dict[str, Any]]:
    """
    Backward-compatible adapter.
    Converts old char-based params into token-based chunking for TXT.
    """
    # Rough conversion factor for Korean-heavy docs.
    max_tokens = max(240, int(max_chars / 2.2))
    min_tokens = max(120, int(min_chunk_chars / 2.2))
    overlap_ratio = max(0.1, min(0.2, overlap / max(1, max_chars)))
    target_tokens = int((min_tokens + max_tokens) / 2)
    return chunk_txt_items(
        lines,
        target_tokens=target_tokens,
        min_tokens=min_tokens,
        max_tokens=max_tokens,
        overlap_ratio=overlap_ratio,
    )
